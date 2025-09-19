import numpy as np
import pandas as pd
from openpyxl import load_workbook
import re

year = "2022"

function_connection = """
1	Állami működési funkciók
F01	Általános közszolgáltatások
F01.a	 Törvényhozó és végrehajtó szervek
F01.b	Pénzügyi és költségvetési tevékenységek és szolgáltatások
F01.c	Külügyek
F01.d	Alapkutatás
F01.e	Műszaki fejlesztés
F01.f	Egyéb általános közösségi szolgáltatások
F02	Védelem
F03	Rendvédelem és közbiztonság
F03.a	Igazságszolgáltatás
F03.b	Rend- és közbiztonság
F03.c	Tűzvédelem
F03.d	Büntetésvégrehajtás
2	Jóléti funkciók
F04	Oktatás
F04.a	Iskolai előkészítés és alapfokú oktatás
F04.b	Középfokú oktatás
F04.c	Felsőfokú oktatás
F04.d	Egyéb oktatás
F05	Egészségügy
F05.a	Kórházi tevékenységek és szolgáltatások
F05.b	Háziorvosi és házi gyermekorvosi ellátás
F05.c	Egyéb rendelői, orvosi, fogorvosi ellátás
F05.d	Közegészségügyi tevékenységek és szolgáltatások
F05.e	Egyéb egészségügy
F06	Társadalombiztosítási, szociális és jóléti szolgáltatások
F06.a	Táppénz, anyasági vagy ideiglenes rokkantsági juttatások
F06.b	Nyugellátások
F06.c	Egyéb társadalombiztosítási ellátások
F06.d	Munkanélküli ellátások 
F06.e	Családi pótlékok és gyermekeknek járó juttatások
F06.f	Egyéb szociális támogatások
F06.g	Jóléti szolgáltatások
F07	Lakásügyek, települési és kommunális szolgáltatások
F08	Szabadidős, kulturális és vallási tevékenységek és szolgáltatások
F08.a	Sport és szabadidős tevékenységek és szolgáltatások
F08.b	Kulturális tevékenységek és szolgáltatások
F08.c	Műsorszórási és kiadói tevékenységek és szolgáltatások
F08.d	Hitéleti tevékenységek
F08.e	Párttevékenységek
F08.f	Egyéb közösségi és szórakoztató tevékenységek
3	Gazdasági funkciók
F09	Tüzelő-, üzemanyag - és energiaellátás
F10	Mező- erdő-, hal- és vadgazdálkodás
F11	Bányászat és ipar
F12	Közlekedés és távközlés
F12.a	Közúti közlekedés
F12.b	Vasúti közlekedés
F12.c	Távközlés
F12.d	Egyéb közlekedés és szállítás
F13	Egyéb gazdasági tevékenységek és szolgáltatások
F13.a	Többcélú fejlesztési programok
F13.b	Egyéb gazdasági tevékenységek és szolgáltatások
F14	Környezetvédelem
4	Államadósság-kezelés
F15	Államadósság kezelés, államháztartás finanszírozása
5	Funkciókba nem sorolható tételek
F16	A főcsoportokba nem sorolható tételek"""

function_dict = {line.split("\t")[0]: line.split("\t")[1] for line in function_connection.strip().split("\n")}

print(function_dict)

# Helper functions that rely on per-year globals (df, fid_names)
def find_name(fid):
    for f, name in fid_names:
        if type(name) != str:
            continue
        if f.startswith(fid+".0"):
            return name.replace("\n", " ").replace("  ", " ")
    print("none for:" +fid)
    return None


def format_fejezet(n, sumcolumn="sum"):
    result = []
    frows = []
    for i, row in df[df["fid"].str.startswith(n)].iterrows():
        fid = row["fid"]
        name = row["name"].replace("\n", " ").replace("  ", " ")
        sum = int(round(row[sumcolumn]*1_000_000))
        if sum < 1:
            continue
        flist = fid.split(".")
        func = str(row["predicted_function"]).strip(":")
        if func == 'nan':
            func = ''
        frows.append({"fid": fid, "name": name, "flist": flist, sumcolumn: sum})
        result.append({"fid": fid, "name": name, sumcolumn: sum, "function": func})

    done = set()
    for level in range(1, 10):
        numbers = [
            ".".join(row["flist"][:level]) for row in frows if len(row["flist"]) > level
        ]
        distinct_numbers = set(numbers)
        for n in distinct_numbers:
            if numbers.count(n) > 1 and n not in done:
                result.append({"fid": n, "name": find_name(n)})
                done.add(n)

    return result


def _build_simple_code_map(items, first_level=False):
    """
    Build a simplified code map for the given items.
    If the items are:
    1.2, 1.2.3, 1.2.4, 1.2.5
    the map should map them to
    01, 02, 03
    respectively
    if the first_level is True the leading zeros will be removed

    more complex mapping:
    1.2, 1.2.3, 1.2.4, 1.3, 1.4
    the map should map them to
    01, 0101, 0102, 03, 04
    to keep the hierarchy
    or if first_level is True
    1, 101, 102, 03, 04

    there may be cases where the hierarchy can be simplified, for example:
    1.2.3, 1.2.3.1, 1.2.3.2, 1.2.4.1
    the map should map them to
    01, 0101, 0102, 02
    to keep the hierarchy
    or if first_level is True
    1, 101, 102, 2

    there may be even cases where more simplification is needed:
    1.2.3.4, 1.3, 1.4
    the map should map them to
    01, 02, 03
    since the long number is there by itself

    returns a dictionary which maps fids to their simplified codes
    """
    fids = [item["fid"] for item in items]
    if not fids:
        return {}

    code_map = {}

    # Group fids by their common parent, even if the parent is not in the list
    # The key for grouping is the parent fid.
    # The value is a list of children fids.
    children_map = {}
    fid_set = set(fids)

    for fid in fids:
        parts = fid.split(".")
        if len(parts) > 1:
            parent = ".".join(parts[:-1])
            # Find the nearest existing ancestor in the fids list
            while parent and parent not in fid_set:
                parent_parts = parent.split(".")
                if len(parent_parts) > 1:
                    parent = ".".join(parent_parts[:-1])
                else:
                    parent = None

            if parent not in children_map:
                children_map[parent] = []
            children_map[parent].append(fid)
        else:  # Root node
            if None not in children_map:
                children_map[None] = []
            children_map[None].append(fid)

    # # Simplify hierarchy: if a node has only one child, its grandchildren become its children
    # for parent_fid, children in list(children_map.items()):
    #     if parent_fid is not None and len(children) == 1:
    #         child_fid = children[0]
    #         grandchildren = children_map.get(child_fid, [])
    #         if grandchildren:
    #             children_map[parent_fid] = grandchildren
    #             children_map[child_fid] = []

    # The fids are pre-sorted, so we can rely on their order within groups.
    # We process level by level.

    # Start with root nodes
    q = children_map.get(None, [])

    for i, fid in enumerate(q):
        index = i + 1
        if first_level:
            code_map[fid] = str(index)
        else:
            code_map[fid] = f"{index:02d}"

    head = 0
    while head < len(q):
        parent_fid = q[head]
        head += 1

        parent_code = code_map.get(parent_fid, "")
        children = children_map.get(parent_fid, [])

        for i, child_fid in enumerate(children):
            index = i + 1
            code_map[child_fid] = f"{parent_code}{index:02d}"
            q.append(child_fid)

    return code_map


def format_items(items, starting="K1", first_level=False):
    code_map = _build_simple_code_map(items, first_level=first_level)
    filtered_items = []
    for item in items:
        fid = item.get("fid")
        code = code_map.get(fid, "")
        if code:
            item["formatted"] = (
                f"{item['name']} ({starting}{code})" if code else item["name"]
            )
            item["budget_id"] = f"{starting}{code}" if code else item["name"]
            filtered_items.append(item)
        else:
            print(f"Warning: No code for fid {fid}!")
    return filtered_items


def extract_budget_id(name):
    match = re.search(r"\((.*?)\)", name)
    return match.group(1) if match else None


def generate_for_year(year: str, income=False) -> pd.DataFrame:
    global df, fid_names  # used by helpers above
    print(f"Generating budget for year: {year}")
    sumcolumn = "income_sum" if income else "sum"

    # Load inputs for the given year
    type_name = "bevétel" if income else "kiadás"
    df_budget = pd.read_excel("budgetatnezes.xlsx", sheet_name=f"budgetdef_{type_name}")
    budgetdef = []
    for i, item in df_budget.iterrows():
        budgetdef.append({"name": item["Megnevezés"], "fid": item["hely"]})

    df = pd.read_excel("manuális címkézés v3.xlsx", year)
    kdf = pd.read_excel("adatok/koltsegvetesek.xlsx", sheet_name=year)

    print(f"sum of df for year {year}:", df[sumcolumn].sum())

    distinct_fejezet = [str(int(f)) for f in kdf["FEJEZET"].dropna().unique().tolist()]

    # Build fid_names from kdf (for find_name)
    fejezet = "0"
    cim = "0"
    alcim = "0"
    jogcim1 = "0"
    jogcim2 = "0"
    fid_names = []
    def valid_num(field):
        return str(field) != "nan" and str(field) and str(field) != "0"
    for index, row in kdf.iterrows():
        if valid_num(row["FEJEZET"]):
            if fejezet != str(row["FEJEZET"]).removesuffix(".0"):
                cim = "0"
                alcim = "0"
                jogcim1 = "0"
                jogcim2 = "0"
            fejezet = str(row["FEJEZET"]).removesuffix(".0")
        if valid_num(row["CIM"]):
            if cim != str(row["CIM"]).removesuffix(".0"):
                alcim = "0"
                jogcim1 = "0"
                jogcim2 = "0"
            cim = str(row["CIM"]).removesuffix(".0")
        if valid_num(row["ALCIM"]):
            if alcim != str(row["ALCIM"]).removesuffix(".0"):
                jogcim1 = "0"
                jogcim2 = "0"
            alcim = str(row["ALCIM"]).removesuffix(".0")
        if valid_num(row["JOGCIM1"]):
            if jogcim1 != str(row["JOGCIM1"]).removesuffix(".0"):
                jogcim2 = "0"
            jogcim1 = str(row["JOGCIM1"]).removesuffix(".0")
        if valid_num(row["JOGCIM2"]):
            jogcim2 = str(row["JOGCIM2"]).removesuffix(".0")
        fid = ".".join([fejezet, cim, alcim, jogcim1, jogcim2])
        fid_names.append((fid, row["MEGNEVEZÉS"]))

    new_budgetdef = []
    fid_to_budget_id = {}
    for row in budgetdef:
        fids = [f.strip() for f in str(row["fid"]).split(",") if f.strip() != "nan"]
        budget_id = extract_budget_id(row["name"])
        if budget_id:
            for fid in fids:
                fid_to_budget_id[fid] = budget_id
            row["budget_id"] = budget_id
        row["formatted"] = row["name"]
        if len(fids) > 1 or not fids:
            new_budgetdef.append(row)
        index_multi = 0
        for fid in fids:
            if fid in distinct_fejezet:
                distinct_fejezet.remove(fid)
            prefix = f"{index_multi + 1:02d}" if len(fids) > 1 else ""
            if fid and fid != "nan":
                fejezet_list = format_fejezet(fid + ".", sumcolumn=sumcolumn)
                fejezet_list = [f for f in fejezet_list if f["fid"]]
                sorted_fejezet = sorted(
                    fejezet_list, key=lambda s: tuple(int(p) for p in s["fid"].split("."))
                )
                formatted_fejezet = format_items(
                    sorted_fejezet,
                    starting=row["budget_id"] + prefix if row.get("budget_id") else "",
                    first_level=(len(row.get("budget_id", "")) == 2),
                )
                if len(formatted_fejezet) >= 1:
                    index_multi += 1
                new_budgetdef.append(
                    {"budget_id": row.get("budget_id"), "formatted": f"{row['name']}"}
                )
                new_budgetdef.extend(formatted_fejezet)

    if distinct_fejezet:
        print(f"!!! Warning: Unmatched fejezet IDs found - {distinct_fejezet}")

    # Deduplicate by budget_id
    seen_budget_ids = set()
    deduplicated_budgetdef = []
    for item in new_budgetdef:
        budget_id = item.get("budget_id")
        if budget_id is None or budget_id not in seen_budget_ids:
            deduplicated_budgetdef.append(item)
            if budget_id is not None:
                seen_budget_ids.add(budget_id)
    new_budgetdef = deduplicated_budgetdef

    totalsum = sum(item.get(sumcolumn, 0) for item in new_budgetdef if item.get(sumcolumn, 0) > 0)

    print(f"Total {type_name} sum for year {year}: {totalsum}")

    # Roll-up sums
    temp_budgetdef = new_budgetdef.copy()
    for row in temp_budgetdef:
        if not row.get(sumcolumn, None):
            row[sumcolumn] = 0
            for item in new_budgetdef:
                if (
                    item.get("budget_id")
                    and row.get("budget_id")
                    and item.get("budget_id").startswith(row.get("budget_id"))
                    and item.get("budget_id") != row.get("budget_id")
                    and len(row["budget_id"]) < len(item.get("budget_id", ""))
                ):
                    row[sumcolumn] += item.get(sumcolumn, 0)
                    if "function" in item:
                        if item["function"] not in row:
                            row[item["function"]] = 0
                        row[item["function"]] += item.get(sumcolumn, 0)
    new_budgetdef = temp_budgetdef
    new_budgetdef = [b for b in new_budgetdef if b.get(sumcolumn, 0) > 0]

    # Trim single-child "01" branches
    trimable_items = [
        b for b in new_budgetdef
        if b.get("budget_id", "").endswith("01")
        and not any([bb for bb in new_budgetdef if bb.get("budget_id").startswith(b["budget_id"][:-2]) and bb.get("budget_id") != b["budget_id"]])
    ]
    skip_items = set()
    while trimable_items:
        item = trimable_items[0]
        item_id = item["budget_id"]
        parent_item = next((bb for bb in new_budgetdef if bb.get("budget_id") == item["budget_id"][:-2]), None)
        parent_id = item["budget_id"][:-2]
        if parent_id in [b.get("budget_id") for b in budgetdef]:
            skip_items.add(item_id)
            trimable_items = [
                b for b in new_budgetdef
                if b.get("budget_id", "").endswith("01")
                and not any([bb for bb in new_budgetdef if bb.get("budget_id") == b["budget_id"][:-2] + '02'])
                and len(b.get("budget_id", "")) > 2
                and b["budget_id"] not in skip_items
            ]
            continue
        if parent_item:
            # print(f"Trimming {item_id} into {parent_id}")
            # print(f"fids: {item.get('fid')} -> {parent_item.get('fid')}")
            new_budgetdef.remove(parent_item)
        for b in new_budgetdef:
            if b.get("budget_id", "").startswith(item_id):
                b["budget_id"] = b["budget_id"].replace(item_id, parent_id)
        trimable_items = [
            b for b in new_budgetdef
            if b.get("budget_id", "").endswith("01")
            and not any([bb for bb in new_budgetdef if bb.get("budget_id") == b["budget_id"][:-2] + '02'])
            and len(b.get("budget_id", "")) > 2
            and b["budget_id"] not in skip_items
        ]

    # Ensure formatted includes budget_id suffix
    for item in new_budgetdef:
        if item.get("budget_id"):
            item["formatted"] = re.sub(r"\(.*?\)$", f"({item['budget_id']})", item.get("formatted", ""))

    for _, row in df.iterrows():
        if row["fid"] not in [b.get("fid") for b in new_budgetdef]:
            pass
            # print(f"Warning: fid {row['fid']} not in budgetdef!")

    # Sum by function
    sum_by_function = {}
    for item in new_budgetdef:
        item[sumcolumn] = item.get(sumcolumn, 0)
        func = item.get("function")
        if func:
            sum_by_function[func] = sum_by_function.get(func, 0) + item.get(sumcolumn, 0)

    for fcode in sum_by_function:
        if fcode not in function_dict:
            print(f"Warning: Unknown function code {fcode} in budgetdef!")

    new_budgetdef.append(
        {
            "formatted": "Bevételek összesen" if income else "Kiadások összesen",
            sumcolumn: totalsum,
        } | sum_by_function
    )

    # pprint.pprint(new_budgetdef[-1])

    df_final = pd.DataFrame(new_budgetdef)

    # Columns and renaming (same as existing)
    function_cols = [
        'F01.a', 'F01.b', 'F01.c', 'F01.d', 'F01.e', 'F01.f',
        'F02',
        'F03.a', 'F03.b', 'F03.c', 'F03.d',
        'F04.a', 'F04.b', 'F04.c', 'F04.d',
        'F05.a', 'F05.b', 'F05.c', 'F05.d', 'F05.e',
        'F06.a', 'F06.b', 'F06.c', 'F06.d', 'F06.e', 'F06.f', 'F06.g',
        'F07',
        'F08.a', 'F08.b', 'F08.c', 'F08.d', 'F08.e', 'F08.f',
        'F09',
        'F10',
        'F11',
        'F12.a', 'F12.b', 'F12.c', 'F12.d',
        'F13.a', 'F13.b',
        'F14',
        'F15',
        'F16',
    ]
    desired_cols = [
        '#', 'Megnevezés', 'Összesen', 'function',
    ] + function_cols

    df_final["#"] = 99
    df_final["Megnevezés"] = df_final["formatted"]
    df_final["Összesen"] = df_final[sumcolumn]

    for col in ["formatted", sumcolumn, "budget_id", "name", "fid"]:
        if col in df_final.columns:
            df_final = df_final.drop(columns=[col])

    df_final = df_final.loc[:, [c for c in df_final.columns if pd.notna(c)]]

    for col in desired_cols:
        if col not in df_final.columns:
            df_final[col] = 0

    df_final = df_final.loc[:, desired_cols]
    numeric_cols = ["Összesen"] + function_cols
    df_final[numeric_cols] = (
        df_final[numeric_cols]
        .apply(pd.to_numeric, errors="coerce")
        .replace([np.inf, -np.inf], 0)
        .fillna(0)
        .astype("int64")
    )

    for col in numeric_cols:
        df_final[col] = df_final[col].apply(lambda x: f"{x:,}")

    code_col_renames = {
        col: f"{list(function_dict.keys()).index(col)+1} {function_dict[col]}"
        for col in df_final.columns if col in function_dict
    }
    df_final = df_final.rename(columns=code_col_renames)

    return df_final


if __name__ == "__main__":
    # Generate sheets for years 2020-2025 in a single Excel file
    years = [str(y) for y in range(2020, 2027)]
    # years = [str(y) for y in [2024]]
    output_path = "budget.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for y in years:
            df_out = generate_for_year(y)
            df_in = generate_for_year(y, income=True)
            df_out.to_excel(writer, index=False, sheet_name=f"{y} KIADÁS")
            df_in.to_excel(writer, index=False, sheet_name=f"{y} BEVÉTEL")


    # Insert title row (A1) into each sheet
    wb = load_workbook(output_path)
    for y in years:
        sheet_name = f"{y} KIADÁS"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.insert_rows(1)
            ws["A1"] = f"{y} KIADÁS"
        sheet_name = f"{y} BEVÉTEL"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.insert_rows(1)
            ws["A1"] = f"{y} BEVÉTEL"
    wb.save(output_path)
